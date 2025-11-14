import json
import os
from typing import Dict, List, Optional

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from e

# Absolute paths (consistent with existing scripts in this repo)
WORKBOOK_PATH = "/Users/divyammakar/workspace/Projects/my-dashboard/whatsapp_workbook.xlsx"
OUTPUT_JSON = "/Users/divyammakar/workspace/Projects/my-dashboard/whatsapp_groups.json"

# Target sheets and form_type mapping
TARGET_SHEETS_FORM_TYPE = {
    "Sakti Team": "shakti",      # Workbook sheet label as provided
    "Shakti Team": "shakti",     # Alias (typo-safe)
    "WTM Group": "wtm",
    "public": "public",
}

# Required headers to extract (exact display keys to keep in JSON objects)
REQUIRED_HEADERS = [
    "Assembly",
    "Group Name",
    "Group Link",
    "Group Members",
    "Admin",
]


def norm(s: str) -> str:
    """Normalize header comparisons: lowercase and collapse whitespace."""
    return " ".join(str(s).strip().lower().split())


def build_sheet_index_map(wb) -> Dict[str, str]:
    """Map normalized sheet name to actual sheet name for robust lookup."""
    index = {}
    for name in wb.sheetnames:
        index[norm(name)] = name
    return index


def find_header_row(ws, required_headers: List[str]) -> Optional[int]:
    """Find the row index (1-based) that contains the headers. Scan first 20 rows."""
    required_norm = {norm(h): h for h in required_headers}
    for row in ws.iter_rows(min_row=1, max_row=20):
        values = [cell.value for cell in row]
        header_cells = {norm(v): i + 1 for i, v in enumerate(values) if v is not None and str(v).strip() != ""}
        if all(rh in header_cells for rh in required_norm):
            return row[0].row
    return None


def get_header_col_map(ws, header_row: int) -> Dict[str, int]:
    """Return mapping from exact header (as in REQUIRED_HEADERS) to column index (1-based)."""
    header_cells = ws[header_row]
    col_map: Dict[str, int] = {}

    # Build a lookup by normalized header text -> col
    found: Dict[str, int] = {}
    for cell in header_cells:
        if cell.value is None:
            continue
        key = norm(str(cell.value))
        found[key] = cell.column

    for header in REQUIRED_HEADERS:
        key = norm(header)
        if key in found:
            col_map[header] = found[key]
        else:
            # Column not present, still include with None to fill empty later
            col_map[header] = None  # type: ignore

    return col_map


def cell_to_string(value) -> str:
    """Convert any cell value to string; return empty string for None."""
    if value is None:
        return ""
    # Keep the exact content as text; avoid trimming Assembly (per requirement)
    return str(value)


def extract_rows(ws, form_type: str) -> List[Dict[str, str]]:
    header_row = find_header_row(ws, REQUIRED_HEADERS)
    if header_row is None:
        print(f"  ⚠️  Could not locate header row for sheet '{ws.title}'. Skipping.")
        return []

    col_map = get_header_col_map(ws, header_row)

    results: List[Dict[str, str]] = []
    # Iterate from the next row after header to the end of sheet
    for r in range(header_row + 1, ws.max_row + 1):
        # Gather values for each required header
        row_obj: Dict[str, str] = {}
        all_empty = True

        for header, col_idx in col_map.items():
            if col_idx is None:
                # Header missing in sheet -> set empty string
                row_obj[header] = ""
                continue
            value = ws.cell(row=r, column=col_idx).value
            text = cell_to_string(value)
            if text != "":
                all_empty = False

            # Special handling: keep Assembly exactly as present (no strip)
            if header == "Assembly":
                row_obj[header] = str(value) if value is not None else ""
            else:
                # For other fields, normalize whitespace lightly (strip only)
                row_obj[header] = text.strip()

        # If the whole row is empty across required columns, skip
        if all_empty:
            continue

        # Add form_type field
        row_obj["form_type"] = form_type

        results.append(row_obj)

    return results


def main():
    if not os.path.exists(WORKBOOK_PATH):
        raise SystemExit(f"Workbook not found at: {WORKBOOK_PATH}")

    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet_index = build_sheet_index_map(wb)

    all_rows: List[Dict[str, str]] = []

    # Resolve each target sheet robustly
    for desired_name, form_type in TARGET_SHEETS_FORM_TYPE.items():
        # Try exact match first, else normalized lookup
        ws = None
        if desired_name in wb.sheetnames:
            ws = wb[desired_name]
        else:
            nm = norm(desired_name)
            actual = sheet_index.get(nm)
            if actual:
                ws = wb[actual]

        if ws is None:
            print(f"  ⚠️  Sheet '{desired_name}' not found. Available: {wb.sheetnames}")
            continue

        print(f"Processing: {ws.title} (form_type={form_type})...")
        rows = extract_rows(ws, form_type)
        print(f"  ✅ Extracted {len(rows)} rows from '{ws.title}'")
        all_rows.extend(rows)

    # Write output JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(all_rows, f, ensure_ascii=False, indent=2)

    print("\n========================================")
    print("Extraction complete")
    print(f"Total rows: {len(all_rows)}")
    print(f"Saved to: {OUTPUT_JSON}")
    if all_rows:
        print("\nSample (first 3 rows):")
        for i, rec in enumerate(all_rows[:3], 1):
            print(f"{i}. {rec}")


if __name__ == "__main__":
    main()
