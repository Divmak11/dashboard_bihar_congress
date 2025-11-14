import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('/Users/divyammakar/workspace/Projects/my-dashboard/workboox.xlsx')

# Define the sheets and their configurations
sheet_configs = [
    {"name": "LAKHISARAI AC", "header_row": 8, "end_row": 75},
    {"name": "HARLAKHI AC", "header_row": 6, "end_row": 21},
    {"name": "WAZIRGANJ AC", "header_row": 8, "end_row": 34},
    {"name": "HARNAUT AC", "header_row": 8, "end_row": 84},
    {"name": "ARWAL AC", "header_row": 6, "end_row": 51},
    {"name": "TIKARI AC", "header_row": 8, "end_row": 38},
    {"name": "PRANPUR AC", "header_row": 6, "end_row": 53},
    {"name": "KISHANGANJ AC", "header_row": 6, "end_row": 34},
    {"name": "AURANGABAD", "header_row": 6, "end_row": 38},
    {"name": "CHIRAYA AC", "header_row": 6, "end_row": 34},
    {"name": "KAHALGAON AC", "header_row": 6, "end_row": 65},
    {"name": "PARO AC", "header_row": 6, "end_row": 28},
    {"name": "JALE AC", "header_row": 6, "end_row": 50},
    {"name": "GOPALGANJ AC", "header_row": 7, "end_row": 28},
    {"name": "ROSERA AC", "header_row": 1, "end_row": 48},
    {"name": "RAJGIR AC", "header_row": 6, "end_row": 48},
]

def normalize_assembly_name(sheet_name):
    """Extract first word and normalize capitalization"""
    # Remove 'AC' suffix and get first word
    first_word = sheet_name.replace(' AC', '').strip().split()[0]
    # Capitalize only first letter, rest lowercase
    return first_word.capitalize()

def extract_data_from_sheet(ws, config):
    """Extract Name and Mobile Number from a sheet"""
    header_row = config["header_row"]
    end_row = config["end_row"]
    assembly = normalize_assembly_name(config["name"])
    
    # Find column indices for Name and Mobile Number
    name_col = None
    mobile_col = None
    
    for cell in ws[header_row]:
        if cell.value:
            value = str(cell.value).strip().lower()  # Convert to lowercase for case-insensitive matching
            if 'name' in value:
                name_col = cell.column
            elif 'mobile' in value or 'number' in value:
                mobile_col = cell.column
    
    if not name_col or not mobile_col:
        print(f"  ⚠️  Could not find Name or Mobile Number columns in {config['name']}")
        return []
    
    data = []
    for row_num in range(header_row + 1, end_row + 1):
        name_cell = ws.cell(row=row_num, column=name_col)
        mobile_cell = ws.cell(row=row_num, column=mobile_col)
        
        name = name_cell.value
        mobile = mobile_cell.value
        
        # Skip empty rows
        if not name and not mobile:
            continue
        
        # Skip rows that don't have both values (as mentioned for ARWAL AC)
        if not name or not mobile:
            continue
        
        # Clean up the values
        name = str(name).strip()
        mobile = str(mobile).strip()
        
        # Skip if values are essentially empty
        if not name or not mobile or name == 'None' or mobile == 'None':
            continue
        
        data.append({
            "name": name,
            "mobile_number": mobile,
            "assembly": assembly
        })
    
    return data

# Extract data from all sheets
print("\n" + "="*60)
print("EXTRACTING DATA FROM SHEETS")
print("="*60 + "\n")

all_data = []
total_records = 0

for config in sheet_configs:
    print(f"Processing: {config['name']}...")
    try:
        ws = wb[config['name']]
        sheet_data = extract_data_from_sheet(ws, config)
        all_data.extend(sheet_data)
        total_records += len(sheet_data)
        print(f"  ✅ Extracted {len(sheet_data)} records from {config['name']}")
    except Exception as e:
        print(f"  ❌ Error processing {config['name']}: {str(e)}")

print(f"\n" + "="*60)
print(f"EXTRACTION COMPLETE")
print(f"Total records extracted: {total_records}")
print("="*60)

# Save to JSON file
output_file = '/Users/divyammakar/workspace/Projects/my-dashboard/extracted_assembly_data.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, ensure_ascii=False)

print(f"\n✅ Data saved to: {output_file}")
print(f"\nSample data (first 3 records):")
for i, record in enumerate(all_data[:3]):
    print(f"\n{i+1}. {record}")
