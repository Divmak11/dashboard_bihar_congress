#!/usr/bin/env python3
"""
Extract SLP Identification data from workbook.xlsx and prepare for Firebase upload.

This script extracts data from three sheets:
- OLD APPLICANT
- SLP APPICANT
- SLP MISSED CALL

For each record, it:
1. Extracts all fields from the sheet
2. Adds 'assembly' field (normalized)
3. Adds 'form_type': 'slp_identification'
4. Adds 'sheet_source' to track which sheet the record came from
5. Outputs to slp_identification_data.json
"""

import openpyxl
import json
from datetime import datetime
import sys

def clean_value(value):
    """Clean cell values for JSON serialization."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Convert numeric phone numbers to strings
        return str(int(value)) if value == int(value) else str(value)
    return str(value).strip() if str(value).strip() else None

def normalize_assembly(assembly_value):
    """Normalize assembly name by removing extra spaces and standardizing format."""
    if not assembly_value:
        return None
    return str(assembly_value).strip()

def extract_sheet_data(sheet, sheet_name):
    """Extract only Name, Mobile Number, and Assembly from a sheet."""
    print(f"\n--- Processing {sheet_name} ---")
    
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        header = clean_value(cell.value)
        if header:
            headers.append(header)
        else:
            headers.append(f"Column_{len(headers) + 1}")
    
    print(f"Headers: {len(headers)} columns")
    print(f"Total rows: {sheet.max_row}")
    
    # Find column indices for required fields
    name_idx = None
    mobile_idx = None
    assembly_idx = None
    
    for idx, header in enumerate(headers):
        if header == 'Name':
            name_idx = idx
        elif header == 'Mobile Number':
            mobile_idx = idx
        elif header == 'Assembly':
            assembly_idx = idx
    
    if name_idx is None or mobile_idx is None or assembly_idx is None:
        print(f"  Error: Could not find required columns (Name, Mobile Number, Assembly)")
        return []
    
    records = []
    skipped_count = 0
    
    # Process data rows (starting from row 2)
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]
        
        # Extract only required fields
        name = clean_value(row[name_idx].value) if name_idx < len(row) else None
        mobile = clean_value(row[mobile_idx].value) if mobile_idx < len(row) else None
        assembly = clean_value(row[assembly_idx].value) if assembly_idx < len(row) else None
        
        # Skip if any required field is missing
        if not name or not mobile or not assembly:
            skipped_count += 1
            continue
        
        # Normalize assembly
        normalized_assembly = normalize_assembly(assembly)
        
        if not normalized_assembly:
            print(f"  Warning: Row {row_idx} has no Assembly value, skipping")
            skipped_count += 1
            continue
        
        # Create minimal record with only required fields
        record = {
            'Name': name,
            'Mobile Number': mobile,
            'assembly': normalized_assembly,
            'form_type': 'slp_identification',
            'sheet_source': sheet_name
        }
        
        records.append(record)
    
    print(f"Extracted: {len(records)} valid records")
    print(f"Skipped: {skipped_count} rows (empty or missing assembly)")
    
    return records

def main():
    """Main extraction function."""
    try:
        print("=" * 60)
        print("SLP Identification Data Extraction")
        print("=" * 60)
        
        # Load workbook
        print("\nLoading workbook.xlsx...")
        wb = openpyxl.load_workbook('workbook.xlsx', data_only=True)
        
        all_records = []
        
        # Extract data from each sheet
        sheets_to_process = [
            ('OLD APPLICANT', 'old_applicant'),
            ('SLP APPICANT', 'slp_applicant'),
            ('SLP MISSED CALL', 'slp_missed_call')
        ]
        
        for sheet_display_name, sheet_source_name in sheets_to_process:
            sheet = wb[sheet_display_name]
            records = extract_sheet_data(sheet, sheet_source_name)
            all_records.extend(records)
        
        # Summary
        print("\n" + "=" * 60)
        print("EXTRACTION SUMMARY")
        print("=" * 60)
        print(f"Total records extracted: {len(all_records)}")
        
        # Count by sheet source
        source_counts = {}
        for record in all_records:
            source = record['sheet_source']
            source_counts[source] = source_counts.get(source, 0) + 1
        
        print("\nRecords by sheet:")
        for source, count in source_counts.items():
            print(f"  {source}: {count}")
        
        # Count by assembly
        assembly_counts = {}
        for record in all_records:
            assembly = record['assembly']
            assembly_counts[assembly] = assembly_counts.get(assembly, 0) + 1
        
        print(f"\nUnique assemblies: {len(assembly_counts)}")
        print(f"Top 5 assemblies:")
        sorted_assemblies = sorted(assembly_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        for assembly, count in sorted_assemblies:
            print(f"  {assembly}: {count} records")
        
        # Write to JSON file
        output_file = 'slp_identification_data.json'
        print(f"\nWriting data to {output_file}...")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Success! Data written to {output_file}")
        print(f"   Total records: {len(all_records)}")
        print(f"   File size: {len(json.dumps(all_records)) / 1024:.2f} KB")
        
        return 0
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
