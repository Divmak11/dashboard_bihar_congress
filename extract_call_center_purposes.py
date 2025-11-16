#!/usr/bin/env python3
"""
Extract Call Center Purposes data from workbook.xlsx

Extracts Name, Mobile Number, and Assembly from these sheets:
- WTM
- PRND
- DONOR
- AGGREGATOR
- DIGITAL MEMBERSHIP 1

Each record gets a 'form_type' field set to the sheet name in lowercase.
Output: call_center_purposes_data.json
"""

import openpyxl
import json
from datetime import datetime
import sys

def clean_value(value):
    """Clean cell values for JSON serialization."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Convert numeric phone numbers to strings
        return str(int(value)) if value == int(value) else str(value)
    return str(value).strip() if str(value).strip() else None

def normalize_assembly(assembly_value):
    """Normalize assembly name by removing extra spaces and standardizing format."""
    if not assembly_value:
        return None
    return str(assembly_value).strip()

def extract_sheet_data(sheet, sheet_name, form_type):
    """Extract only Name, Mobile Number, and Assembly from a sheet."""
    print(f"\n--- Processing {sheet_name} ---")
    
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        header = clean_value(cell.value)
        if header:
            headers.append(header)
        else:
            headers.append(f"Column_{len(headers) + 1}")
    
    print(f"Headers: {len(headers)} columns")
    print(f"Total rows: {sheet.max_row}")
    
    # Find column indices for required fields
    name_idx = None
    mobile_idx = None
    assembly_idx = None
    
    for idx, header in enumerate(headers):
        if header == 'Name':
            name_idx = idx
        elif header == 'Mobile Number':
            mobile_idx = idx
        elif header == 'Assembly':
            assembly_idx = idx
    
    # Flag if columns are missing
    missing_columns = []
    if name_idx is None:
        missing_columns.append('Name')
    if mobile_idx is None:
        missing_columns.append('Mobile Number')
    if assembly_idx is None:
        missing_columns.append('Assembly')
    
    if missing_columns:
        print(f"  ⚠️  WARNING: Missing columns in '{sheet_name}': {', '.join(missing_columns)}")
        print(f"  Available headers: {headers[:10]}...")
        return []
    
    records = []
    skipped_count = 0
    
    # Process data rows (starting from row 2)
    for row_idx in range(2, sheet.max_row + 1):
        row = sheet[row_idx]
        
        # Extract only required fields
        name = clean_value(row[name_idx].value) if name_idx < len(row) else None
        mobile = clean_value(row[mobile_idx].value) if mobile_idx < len(row) else None
        assembly = clean_value(row[assembly_idx].value) if assembly_idx < len(row) else None
        
        # Skip if any required field is missing
        if not name or not mobile or not assembly:
            skipped_count += 1
            continue
        
        # Normalize assembly
        normalized_assembly = normalize_assembly(assembly)
        
        if not normalized_assembly:
            skipped_count += 1
            continue
        
        # Create minimal record with only required fields
        record = {
            'Name': name,
            'Mobile Number': mobile,
            'assembly': normalized_assembly,
            'form_type': form_type
        }
        
        records.append(record)
    
    print(f"✅ Extracted: {len(records)} valid records")
    if skipped_count > 0:
        print(f"   Skipped: {skipped_count} rows (empty or missing required fields)")
    
    return records

def main():
    """Main extraction function."""
    try:
        print("=" * 60)
        print("Call Center Purposes Data Extraction")
        print("=" * 60)
        
        # Load workbook
        print("\nLoading workbook.xlsx...")
        wb = openpyxl.load_workbook('workbook.xlsx', data_only=True)
        
        # List all available sheets for debugging
        print(f"\nAvailable sheets in workbook: {wb.sheetnames}")
        
        all_records = []
        
        # Define sheets to process with their form_type values
        sheets_to_process = [
            ('WTM', 'wtm'),
            ('PRND', 'prnd'),
            ('DONOR', 'donor'),
            ('AGGREGATOR', 'aggregator'),
            ('DIGITAL MEMBERSHIP 1', 'digital membership 1'),
        ]
        
        errors = []
        
        for sheet_display_name, form_type in sheets_to_process:
            try:
                if sheet_display_name not in wb.sheetnames:
                    error_msg = f"Sheet '{sheet_display_name}' not found in workbook"
                    print(f"\n❌ ERROR: {error_msg}")
                    errors.append(error_msg)
                    continue
                
                sheet = wb[sheet_display_name]
                records = extract_sheet_data(sheet, sheet_display_name, form_type)
                all_records.extend(records)
            except Exception as e:
                error_msg = f"Error processing sheet '{sheet_display_name}': {str(e)}"
                print(f"\n❌ ERROR: {error_msg}")
                errors.append(error_msg)
        
        # Summary
        print("\n" + "=" * 60)
        print("EXTRACTION SUMMARY")
        print("=" * 60)
        print(f"Total records extracted: {len(all_records)}")
        
        if errors:
            print(f"\n⚠️  Errors encountered: {len(errors)}")
            for error in errors:
                print(f"  - {error}")
        
        # Count by form_type
        form_type_counts = {}
        for record in all_records:
            form_type = record['form_type']
            form_type_counts[form_type] = form_type_counts.get(form_type, 0) + 1
        
        print("\nRecords by form type:")
        for form_type, count in sorted(form_type_counts.items()):
            print(f"  {form_type}: {count}")
        
        # Count by assembly
        assembly_counts = {}
        for record in all_records:
            assembly = record['assembly']
            assembly_counts[assembly] = assembly_counts.get(assembly, 0) + 1
        
        print(f"\nUnique assemblies: {len(assembly_counts)}")
        print(f"Top 5 assemblies:")
        sorted_assemblies = sorted(assembly_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        for assembly, count in sorted_assemblies:
            print(f"  {assembly}: {count} records")
        
        # Write to JSON file
        output_file = 'call_center_purposes_data.json'
        print(f"\nWriting data to {output_file}...")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(all_records, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Success! Data written to {output_file}")
        print(f"   Total records: {len(all_records)}")
        print(f"   File size: {len(json.dumps(all_records)) / 1024:.2f} KB")
        
        if errors:
            print(f"\n⚠️  Note: Some sheets had errors. Please review above.")
            return 1
        
        return 0
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
